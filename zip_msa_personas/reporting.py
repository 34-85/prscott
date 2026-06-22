"""Build a formatted, multi-tab Excel workbook from scored ZIP-level personas.

Reusable for both the preview and the official run. Given a per-ZIP frame with
the persona share columns, it writes:

  * Read me               -- what the file is, column legend, disclosures
  * All ZIPs              -- every ZIP: top persona, confidence, basis,
                             per-segment % and index-vs-national (color-scaled)
  * By State             -- persona mix rolled up to state
  * Footprint by basis   -- for each persona, how much of its footprint is
                            survey data vs estimated (the credibility view)
  * Top <N> - <persona>  -- one ranked target-list tab per segment

Index = a ZIP's segment share / the national average share x 100
(100 = national average; 150 = over-indexes 1.5x).
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd

BASIS_ORDER = [
    "Survey data (your NPOS)",
    "Estimated — similar nearby ZIPs",
    "Estimated — broad, low confidence (disclosed)",
]


def _sheet_name(prefix: str, persona: str) -> str:
    name = f"{prefix} {persona}"
    return name[:31]  # Excel hard limit


def build_workbook(
    zips: pd.DataFrame,
    segments: list[str],
    out_path: str | Path,
    *,
    top_n: int = 100,
    title: str = "APPA pet-owner personas by ZIP",
    data_vintage: str = "",
    methodology_version: str = "",
    is_preview: bool = True,
) -> Path:
    """Write the workbook.

    ``zips`` must contain: ZIP, City, State, Top persona, Confidence, Basis, and
    one share column per segment named exactly as the segment (values 0..1).
    """
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    national = {s: zips[s].mean() for s in segments}

    # ---- assemble the All ZIPs sheet (shares -> % and index columns) ----
    allz = zips[["ZIP", "City", "State", "Top persona", "Confidence", "Basis"]].copy()
    for s in segments:
        allz[f"{s} %"] = (zips[s] * 100).round(1)
    for s in segments:
        allz[f"{s} idx"] = (zips[s] / national[s] * 100).round(0).astype("Int64")
    allz = allz.sort_values(["State", "ZIP"]).reset_index(drop=True)

    # ---- By State ----
    state = (
        zips.groupby("State")[segments].mean().mul(100).round(1).reset_index()
    )

    # ---- Footprint by basis (per dominant persona) ----
    fp = (
        zips.assign(n=1)
        .pivot_table(index="Top persona", columns="Basis", values="n", aggfunc="sum", fill_value=0)
        .reindex(columns=BASIS_ORDER, fill_value=0)
    )
    fp["Total ZIPs"] = fp.sum(axis=1)
    fp["% survey-backed"] = (fp[BASIS_ORDER[0]] / fp["Total ZIPs"] * 100).round(1)
    fp = fp.reset_index()

    # ---- Read me ----
    legend = [
        [title, ""],
        ["Generated", date.today().isoformat()],
        ["Status", "PREVIEW — geographic smoothing + state priors" if is_preview
         else "Official — Census demographics + MSA/DMA"],
        ["Data vintage", data_vintage],
        ["Methodology version", methodology_version],
        ["", ""],
        ["Tab", "What it shows"],
        ["All ZIPs", "Every US ZIP: top persona, confidence, basis, full 7-segment % and index"],
        ["By State", "Average persona mix (%) per state"],
        ["Footprint by basis", "For each persona, how many ZIPs are survey vs estimated"],
        [f"Top {top_n} - <persona>", "Ranked target list of the ZIPs that most over-index on each segment"],
        ["", ""],
        ["Column legend", ""],
        ["Confidence", "0-1; how much to trust the top persona for this ZIP"],
        ["Basis", "Survey data = your NPOS; Estimated = modeled (see disclosure)"],
        ["<segment> %", "Share of that segment among pet owners in the ZIP"],
        ["<segment> idx", "Index vs US average: 100 = average, 150 = over-indexes 1.5x"],
        ["", ""],
        ["Disclosure", "Estimated ZIPs are modeled, not surveyed. 'Broad, low confidence' "
                       "ZIPs resemble no surveyed area and use a regional/national baseline."],
    ]
    readme = pd.DataFrame(legend)

    # ---- write everything ----
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        readme.to_excel(xl, sheet_name="Read me", header=False, index=False)
        allz.to_excel(xl, sheet_name="All ZIPs", index=False)
        state.to_excel(xl, sheet_name="By State", index=False)
        fp.to_excel(xl, sheet_name="Footprint by basis", index=False)
        for s in segments:
            cols = ["ZIP", "City", "State", f"{s} %", f"{s} idx", "Confidence", "Basis"]
            top = allz.sort_values(f"{s} %", ascending=False).head(top_n)[cols].reset_index(drop=True)
            top.to_excel(xl, sheet_name=_sheet_name(f"Top{top_n}", s), index=False)

        wb = xl.book
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        for ws in wb.worksheets:
            if ws.title == "Read me":
                ws.column_dimensions["A"].width = 26
                ws.column_dimensions["B"].width = 95
                ws["A1"].font = Font(bold=True, size=14)
                continue
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            # color-scale any 'idx' columns: red (low) -> white (100) -> green (high)
            for j, name in enumerate(c.value for c in ws[1]):
                if isinstance(name, str) and name.endswith("idx"):
                    col = get_column_letter(j + 1)
                    rng = f"{col}2:{col}{ws.max_row}"
                    ws.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="num", start_value=0, start_color="F8696B",
                            mid_type="num", mid_value=100, mid_color="FFFFFF",
                            end_type="num", end_value=200, end_color="63BE7B",
                        ),
                    )
    return out_path


__all__ = ["build_workbook"]
